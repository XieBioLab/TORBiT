import csv
from collections import Counter
import os
from openpyxl import load_workbook, Workbook
from collections import defaultdict
def analyze_tcr_usage_to_excel(csv_file, output_excel):
    alpha_v_gene_count = Counter()
    alpha_j_gene_count = Counter()
    beta_v_gene_count = Counter()
    beta_j_gene_count = Counter()
    
    with open(csv_file, 'r', encoding='utf-8-sig') as file: 
        reader = csv.DictReader(file, delimiter='\t')  
        
        for row in reader:
            try:
                chain_type = row["Chain_type"].strip()  
                v_gene = row["V"].strip()
                j_gene = row["J"].strip()
                
                if chain_type == "VA":  
                    for v in v_gene.split(","):
                        alpha_v_gene_count[v.strip()] += 1
                    for j in j_gene.split(","):
                        alpha_j_gene_count[j.strip()] += 1
                elif chain_type == "VB":  
                    for v in v_gene.split(","):
                        beta_v_gene_count[v.strip()] += 1
                    for j in j_gene.split(","):
                        beta_j_gene_count[j.strip()] += 1
            except KeyError as e:
                print(f"Row {reader.line_num} is missing key column: {e}")
                continue
    
    wb = Workbook()
    
    sheets_data = [
        ("αV", alpha_v_gene_count),
        ("αJ", alpha_j_gene_count),
        ("βV", beta_v_gene_count),
        ("βJ", beta_j_gene_count)
    ]
    
    for sheet_name, counter in sheets_data:
        sheet = wb.create_sheet(sheet_name)
        sheet.append([f"{sheet_name[0]} Gene", "Count"])
        for gene, count in counter.most_common(): 
            sheet.append([gene, count])
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    wb.save(output_excel)
    print(f"Statistical results have been saved to {output_excel}")

def merge_excel_files(main_dir, output_file):
    sub_dirs = ["P01-T-I", "P01-T-II", "P01-T-III", "P01-T-IV"]
    sheet_names = ["αV", "αJ", "βV", "βJ"]
    
    merged_wb = Workbook()
    while merged_wb.sheetnames:
        merged_wb.remove(merged_wb[merged_wb.sheetnames[0]])
    
    sheet_data = {name: defaultdict(dict) for name in sheet_names}
    sample_names = [d.split("-")[-1] for d in sub_dirs] 
    
    for sheet_name in sheet_names:
        for sample_idx, sub_dir in enumerate(sub_dirs):
            file_path = os.path.join(main_dir, sub_dir, "out.xlsx")
            
            if not os.path.exists(file_path):
                print(f"Warning: File {file_path} does not exist, skipping")
                continue
            
            try:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    print(f"Warning: Sheet {sheet_name} does not exist in {file_path}, skipping")
                    continue
                
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):  
                    if len(row) >= 2 and row[0] is not None:
                        gene_name = str(row[0]).strip()
                        count = row[1] if row[1] is not None else 0
                        sheet_data[sheet_name][gene_name][sample_idx] = count
            except Exception as e:
                print(f"Error occurred while processing {sheet_name} in {file_path}: {e}")
                continue
    
    for sheet_name in sheet_names:
        merged_ws = merged_wb.create_sheet(sheet_name)
        
        header = ["Gene"]
        for name in sample_names:
            header.append(name)
            header.append("")  
        merged_ws.append(header[:-1])  
        
        for gene in sorted(sheet_data[sheet_name].keys()):
            row_data = [gene]
            for sample_idx in range(len(sub_dirs)):
                if sample_idx in sheet_data[sheet_name][gene]:
                    row_data.append(sheet_data[sheet_name][gene][sample_idx])
                else:
                    row_data.append(0)  
                row_data.append("")  
            merged_ws.append(row_data[:-1])
    
    merged_wb.save(output_file)
    print(f"所有文件已按sheet合并到 {output_file}，相同基因已对齐")
