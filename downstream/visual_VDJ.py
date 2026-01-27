import argparse
import pandas as pd
import matplotlib.pyplot as plt
import csv
from collections import Counter
import openpyxl


def count_gene_usage(csv_file, output_excel):
    """Count V and J gene usage from TCR sequencing data."""
    
    # Initialize counters
    alpha_v_gene_count = Counter()
    alpha_j_gene_count = Counter()
    beta_v_gene_count = Counter()
    beta_j_gene_count = Counter()

    # Read CSV file
    with open(csv_file, "r") as file:
        reader = csv.DictReader(file, delimiter="\t")  # Specify tab as delimiter
        for row in reader:
            # Check if Chain_type exists
            if "Chain_type" not in row:
                print("Warning: Missing 'Chain_type' in row:", row)
                continue  # Skip abnormal rows

            chain_type = row["Chain_type"]
            v_gene = row["V"]
            j_gene = row["J"]

            # Check if Chain_type is VA or VB
            if chain_type not in ["VA", "VB"]:
                print(f"Warning: Unknown chain type '{chain_type}' in row:", row)
                continue  # Skip unknown chain types

            # Count V and J gene usage based on chain type
            if chain_type == "VA":
                for v in v_gene.split(","):
                    alpha_v_gene_count[v.strip()] += 1
                for j in j_gene.split(","):
                    alpha_j_gene_count[j.strip()] += 1
            elif chain_type == "VB":
                for v in v_gene.split(","):
                    beta_v_gene_count[v.strip()] += 1
                for j in j_gene.split(","):
                    beta_j_gene_count[j.strip()] += 1

    # Create Excel workbook
    wb = openpyxl.Workbook()

    # Add αV sheet
    alpha_v_sheet = wb.create_sheet("αV")
    alpha_v_sheet.append(["V Gene", "Count"])
    for v_gene, count in alpha_v_gene_count.items():
        alpha_v_sheet.append([v_gene, count])

    # Add αJ sheet
    alpha_j_sheet = wb.create_sheet("αJ")
    alpha_j_sheet.append(["J Gene", "Count"])
    for j_gene, count in alpha_j_gene_count.items():
        alpha_j_sheet.append([j_gene, count])

    # Add βV sheet
    beta_v_sheet = wb.create_sheet("βV")
    beta_v_sheet.append(["V Gene", "Count"])
    for v_gene, count in beta_v_gene_count.items():
        beta_v_sheet.append([v_gene, count])

    # Add βJ sheet
    beta_j_sheet = wb.create_sheet("βJ")
    beta_j_sheet.append(["J Gene", "Count"])
    for j_gene, count in beta_j_gene_count.items():
        beta_j_sheet.append([j_gene, count])

    # Remove default worksheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Save Excel file
    wb.save(output_excel)
    print(f"Results saved to: {output_excel}")


def all_genes_plot(input_excel, output_plot):
    """Generate a combined plot showing all gene usage statistics."""
    
    # Read data from Excel file
    alpha_v_usage = pd.read_excel(input_excel, sheet_name="αV")
    alpha_j_usage = pd.read_excel(input_excel, sheet_name="αJ")
    beta_v_usage = pd.read_excel(input_excel, sheet_name="βV")
    beta_j_usage = pd.read_excel(input_excel, sheet_name="βJ")

    # Ensure consistent data format
    alpha_v_usage["V Gene"] = alpha_v_usage["V Gene"].astype(str)
    alpha_j_usage["J Gene"] = alpha_j_usage["J Gene"].astype(str)
    beta_v_usage["V Gene"] = beta_v_usage["V Gene"].astype(str)
    beta_j_usage["J Gene"] = beta_j_usage["J Gene"].astype(str)

    # Create a figure with four subplots (2 rows, 2 columns)
    fig, axs = plt.subplots(2, 2, figsize=(16, 12))

    # Plot αV gene usage bar chart
    axs[0, 0].bar(alpha_v_usage["V Gene"], alpha_v_usage["Count"], color="blue")
    axs[0, 0].set_xlabel("V Gene")
    axs[0, 0].set_ylabel("Count")
    axs[0, 0].set_title("αV Gene Usage")
    axs[0, 0].set_xticks(range(len(alpha_v_usage["V Gene"])))
    axs[0, 0].set_xticklabels(alpha_v_usage["V Gene"], rotation=90, 
                               horizontalalignment="right", fontsize=5)

    # Plot αJ gene usage bar chart
    axs[0, 1].bar(alpha_j_usage["J Gene"], alpha_j_usage["Count"], color="green")
    axs[0, 1].set_xlabel("J Gene")
    axs[0, 1].set_ylabel("Count")
    axs[0, 1].set_title("αJ Gene Usage")
    axs[0, 1].set_xticks(range(len(alpha_j_usage["J Gene"])))
    axs[0, 1].set_xticklabels(alpha_j_usage["J Gene"], rotation=90, 
                               horizontalalignment="right", fontsize=6)

    # Plot βV gene usage bar chart
    axs[1, 0].bar(beta_v_usage["V Gene"], beta_v_usage["Count"], color="red")
    axs[1, 0].set_xlabel("V Gene")
    axs[1, 0].set_ylabel("Count")
    axs[1, 0].set_title("βV Gene Usage")
    axs[1, 0].set_xticks(range(len(beta_v_usage["V Gene"])))
    axs[1, 0].set_xticklabels(beta_v_usage["V Gene"], rotation=90, 
                               horizontalalignment="right", fontsize=5)

    # Plot βJ gene usage bar chart
    axs[1, 1].bar(beta_j_usage["J Gene"], beta_j_usage["Count"], color="purple")
    axs[1, 1].set_xlabel("J Gene")
    axs[1, 1].set_ylabel("Count")
    axs[1, 1].set_title("βJ Gene Usage")
    axs[1, 1].set_xticks(range(len(beta_j_usage["J Gene"])))
    axs[1, 1].set_xticklabels(beta_j_usage["J Gene"], rotation=90, 
                               horizontalalignment="right", fontsize=10)

    # Adjust subplot spacing
    plt.tight_layout(pad=3.0)
    plt.savefig(output_plot)
    print(f"Figure saved to: {output_plot}")


def solo_gene_plot(input_excel, output_plot_prefix):
    """Generate individual plots for each gene type."""
    
    # Read data from Excel file
    alpha_v_usage = pd.read_excel(input_excel, sheet_name="αV")
    alpha_j_usage = pd.read_excel(input_excel, sheet_name="αJ")
    beta_v_usage = pd.read_excel(input_excel, sheet_name="βV")
    beta_j_usage = pd.read_excel(input_excel, sheet_name="βJ")

    # Ensure consistent data format
    alpha_v_usage["V Gene"] = alpha_v_usage["V Gene"].astype(str)
    alpha_j_usage["J Gene"] = alpha_j_usage["J Gene"].astype(str)
    beta_v_usage["V Gene"] = beta_v_usage["V Gene"].astype(str)
    beta_j_usage["J Gene"] = beta_j_usage["J Gene"].astype(str)

    # Plot αV gene usage bar chart
    plt.figure(figsize=(10, 6))
    plt.bar(alpha_v_usage["V Gene"], alpha_v_usage["Count"], color="blue", width=0.8)
    # Adjust x-axis range to bring first bar closer to y-axis
    plt.xlim(-0.8, len(alpha_v_usage["V Gene"]) - 0.1)
    plt.xlabel("V Gene")
    plt.ylabel("Count")
    plt.title("αV Gene Usage")
    plt.xticks(rotation=90, horizontalalignment="right", fontsize=7)
    plt.tight_layout()
    plt.savefig(f"{output_plot_prefix}_alpha_v.png")
    plt.close()

    # Plot αJ gene usage bar chart
    plt.figure(figsize=(10, 6))
    plt.bar(alpha_j_usage["J Gene"], alpha_j_usage["Count"], color="green", width=0.8)
    plt.xlabel("J Gene")
    plt.ylabel("Count")
    plt.title("αJ Gene Usage")
    plt.xticks(rotation=90, horizontalalignment="right", fontsize=7)
    plt.tight_layout()
    plt.savefig(f"{output_plot_prefix}_alpha_j.png")
    plt.close()

    # Plot βV gene usage bar chart
    plt.figure(figsize=(10, 6))
    plt.bar(beta_v_usage["V Gene"], beta_v_usage["Count"], color="red", width=0.8)
    plt.xlim(-0.8, len(beta_v_usage["V Gene"]) - 0.1)
    plt.xlabel("V Gene")
    plt.ylabel("Count")
    plt.title("βV Gene Usage")
    plt.xticks(rotation=90, horizontalalignment="right", fontsize=7)
    plt.tight_layout()
    plt.savefig(f"{output_plot_prefix}_beta_v.png")
    plt.close()

    # Plot βJ gene usage bar chart
    plt.figure(figsize=(10, 6))
    plt.bar(beta_j_usage["J Gene"], beta_j_usage["Count"], color="purple")
    plt.xlabel("J Gene")
    plt.ylabel("Count")
    plt.title("βJ Gene Usage")
    plt.xticks(rotation=90, horizontalalignment="right", fontsize=10)
    plt.tight_layout()
    plt.savefig(f"{output_plot_prefix}_beta_j.png")
    plt.close()

    print(f"Figures saved with prefix: {output_plot_prefix}")
