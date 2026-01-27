import csv
from collections import Counter
from openpyxl import Workbook

def analyze_tcr_usage_to_excel(csv_file, output_excel):
    alpha_v_gene_count = Counter()
    alpha_j_gene_count = Counter()
    beta_v_gene_count = Counter()
    beta_j_gene_count = Counter()
    
    with open(csv_file, 'r') as file:
        reader = csv.DictReader(file)
        for row in reader:
            chain_type = row["Chain_type"]
            v_gene = row["V"]
            j_gene = row["J"]
            
            if chain_type == "VA":  
                for v in v_gene.split(","):
                    alpha_v_gene_count[v.strip()] += 1
                for j in j_gene.split(","):
                    alpha_j_gene_count[j.strip()] += 1
            elif chain_type == "VB": 
                for v in v_gene.split(","):
                    beta_v_gene_count[v.strip()] += 1
                for j in j_gene.split(","):
                    beta_j_gene_count[j.strip()] += 1
    
    wb = Workbook()
    
    alpha_v_sheet = wb.create_sheet("αV")
    alpha_v_sheet.append(["V Gene", "Count"])
    for v_gene, count in alpha_v_gene_count.items():
        alpha_v_sheet.append([v_gene, count])
    
    alpha_j_sheet = wb.create_sheet("αJ")
    alpha_j_sheet.append(["J Gene", "Count"])
    for j_gene, count in alpha_j_gene_count.items():
        alpha_j_sheet.append([j_gene, count])
    
    beta_v_sheet = wb.create_sheet("βV")
    beta_v_sheet.append(["V Gene", "Count"])
    for v_gene, count in beta_v_gene_count.items():
        beta_v_sheet.append([v_gene, count])
    
    beta_j_sheet = wb.create_sheet("βJ")
    beta_j_sheet.append(["J Gene", "Count"])
    for j_gene, count in beta_j_gene_count.items():
        beta_j_sheet.append([j_gene, count])
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    wb.save(output_excel)
    print(f"Statistics have been saved to{output_excel}")


